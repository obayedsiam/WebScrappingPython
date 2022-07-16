import openpyxl
import requests
import selenium
from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import pandas as pd


def getBookUrlFromWriterUrl(options, driver):
    # Collecting Books URL
    booksLinkList = []
    basicUrl = "https://www.rokomari.com"
    empltyBookUrl = 0
    c = {'Books Link': booksLinkList}
    linkUrlFile = openpyxl.load_workbook('H:\Programming\pyCharm\WebScrappingPython\AllWriterValidUrl_Part_1.xlsx',data_only=True)
    # linkUrlFile = openpyxl.load_workbook(r'C:\Users\EATL\PycharmProjects\WebScapping\AllWriterValidUrl_Part_1.xlsx',data_only=True)
    sheet_obj = linkUrlFile.active
    m_row = sheet_obj.max_row
    # print(m_row)

    file = open("bookUrlNumber.txt", "r")
    readNumber = file.read()
    writer_url_number = int(readNumber)
    file.close()
    # writer_url_number = 18912
    for i in range(writer_url_number, m_row):
        cell_obj = sheet_obj.cell(row=i, column=2)
        linkAddress = cell_obj.value
        file = open("nextUrl.txt", "r")
        nextUrl = file.read()
        file.close()

        if len(nextUrl) > 0:
            linkAddress = nextUrl

        isNextPresent = True
        while isNextPresent:
            print("\nWriter Url Link : ", (i - 2))
            print(linkAddress)
            driver.get(linkAddress)
            content = driver.page_source
            soup = BeautifulSoup(content, features="html.parser")
            LinkListDiv = soup.find_all('div', attrs={'class': 'book-list-wrapper'})
            print("Book List from the Writer Url : ")
            # print(LinkListDiv)
            if len(LinkListDiv) > 0:
                # print("Entered into if cond")
                for listTemp in LinkListDiv:
                    aTag = listTemp.find('a')
                    href = aTag.get('href')
                    print(href)
                    booksLinkList.append(basicUrl+href)
                    url = basicUrl + href
                    myFileName = r'H:\Programming\pyCharm\WebScrappingPython\BookList.xlsx'
                    # myFileName = r'C:\Users\EATL\PycharmProjects\WebScapping\BookList.xlsx'
                    try:
                        wb = load_workbook(filename=myFileName)
                    except:
                        print("Could not load file")

                    ws = wb['Sheet1']
                    newRowLocation = ws.max_row + 1
                    ws.cell(column=2, row=newRowLocation, value=href)
                    wb.save(filename=myFileName)
                    wb.close()
                    try:
                        my_element = driver.find_element_by_xpath("//a[text()='Next']")
                        linkAddress = my_element.get_attribute('href')
                        file = open("nextUrl.txt", "w")
                        file.write(linkAddress)
                        file.close()

                    except NoSuchElementException:
                        isNextPresent = False
                        file = open("bookUrlNumber.txt", "w")
                        file.write(str(i + 1))
                        file.close()
                        # print("Empty Book Url Number :", empltyBookUrl)
                        # print("End of Writer Book List")

                        file = open("nextUrl.txt", "w")
                        file.write("")
                        file.close()
            else:
                isNextPresent = False
                file = open("bookUrlNumber.txt", "w")
                file.write(str(i + 1))
                file.close()

                file = open("nextUrl.txt", "w")
                file.write("")
                file.close()
                # continue
                # i = i + 1
                # cell_obj = sheet_obj.cell(row=i, column=2)
                # linkAddress = cell_obj.value
                # empltyBookUrl = empltyBookUrl + 1


        # Inserting into Excel File
        # df = pd.DataFrame.from_dict(c, orient='index')
        # df = df.transpose()
        # df.to_excel(r'H:\Programming\pyCharm\WebScrappingPython\AllBook.xlsx', index=True, encoding='utf-8')
        # df.to_excel(r'C:\Users\EATL\PycharmProjects\WebScapping\AllBook.xlsx', index=True, encoding='utf-8')

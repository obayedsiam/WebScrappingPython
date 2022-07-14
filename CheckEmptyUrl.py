import openpyxl
from openpyxl import load_workbook
import requests
import selenium
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import pandas as pd

writerUrlNumber222 = 63134


def check_empty_url(options, driver):
    # Collecting Books URL
    file = open("writerUrlNumber.txt", "r")
    readNumber = file.read()
    writer_url_number = int(readNumber)
    file.close()

    validUrlList = []
    basicUrl = "https://www.rokomari.com"
    emptyWriterUrl = 0
    d = {'Writer Valid Url': validUrlList}
    linkUrlFile = openpyxl.load_workbook(r'H:\Programming\pyCharm\WebScrappingPython\AllWriter.xlsx', data_only=True)
    #linkUrlFile = openpyxl.load_workbook(r'C:\Users\EATL\PycharmProjects\WebScapping\AllWriter.xlsx', data_only=True)
    sheet_obj = linkUrlFile.active
    m_row = sheet_obj.max_row
    m_row = 78977
    print(m_row)

    for i in range(writer_url_number, m_row):
        cell_obj = sheet_obj.cell(row=i, column=2)
        linkAddress = cell_obj.value
        isNextPresent = True

        print("Writer Url Link : ", (i - 2))
        print(linkAddress)
        driver.get(linkAddress)
        content = driver.page_source
        soup = BeautifulSoup(content, features="html.parser")
        LinkListDiv = soup.find_all('div', attrs={'class': 'book-list-wrapper'})

        # print("Book List from the Writer Url : ")
        # print(LinkListDiv)
        if len(LinkListDiv) > 0:
            # print("Entered into if cond")
            validUrlList.append(linkAddress)
            myFileName = r'H:\Programming\pyCharm\WebScrappingPython\AllWriterValidUrl2.xlsx'
            #myFileName = r'C:\Users\EATL\PycharmProjects\WebScapping\AllWriterValidUrl2.xlsx'

            wb = load_workbook(filename=myFileName)
            ws = wb['Sheet1']
            newRowLocation = ws.max_row + 1
            ws.cell(column=2, row=newRowLocation, value=linkAddress)
            wb.save(filename=myFileName)
            wb.close()
        else:
            # i = i+1
            #         cell_obj = sheet_obj.cell(row=i, column=2)
            #         linkAddress = cell_obj.value
            emptyWriterUrl = emptyWriterUrl + 1

        try:
            file = open("writerUrlNumber.txt", "w")
            file.write(str(i + 1))
            file.close()
        except NameError:
            print(NameError)

        print("Empty Book Url Number :", emptyWriterUrl)
        # Inserting into Excel File
        # df = pd.DataFrame.from_dict(d, orient='index')
        # df = df.transpose()
        # df.to_excel(r'H:\Pro    gramming\pyCharm\WebScrappingPython\AllWriterValidUrl.xlsx', index=True, encoding='utf-8')
        # df.to_excel(r'C:\Users\EATL\PycharmProjects\WebScapping\AllWriterValidUrl.xlsx', index=True, encoding='utf-8')
